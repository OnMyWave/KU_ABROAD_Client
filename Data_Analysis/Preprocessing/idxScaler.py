from sklearn.preprocessing import MinMaxScaler
import pandas as pd
import openpyxl
from math import isnan

def calc(x):
    if not isnan(x[0]):
        return int(x[0]) + 50

def idxScaler(df,col_name):
    idx = df[col_name][1:].values.reshape(-1,1) # Crime-index
    scaler = MinMaxScaler([0,50])
    idx = list(map(calc , scaler.fit_transform(idx)))
    return idx

def quantileEncoding(df,col_name):
    df = df[col_name][1:]
    quan = [df.quantile(q = i/4) for i in range(1,4)]
    df = list(df)
    for idx in range(len(df)):
        if df[idx] < quan[0] :
            df[idx] = 1
        elif df[idx] < quan[1] :
            df[idx] = 2
        elif df[idx] < quan[2] :
            df[idx] = 3
        else:
            df[idx] = 4
    return df

if __name__ == "__main__":
    univ = pd.read_excel('/Users/onmywave/Desktop/Github/Datathon/Data_Analysis/univ_data.xlsx')
    
    idxes = {
        'living_idx' : '생활 정보',
        'crime_idx' : 'Unnamed: 29',
        'gvi_idx' : 'Unnamed: 35',
        'sdg_idx' : 'Unnamed: 36',
        }
    for idx_name, col in {'gvi_idx' : 'Unnamed: 35',
        'sdg_idx' : 'Unnamed: 36'}.items():
        idxes[idx_name] = idxScaler(univ,col)
    for idx_name, col in {'living_idx' : '생활 정보',
        'crime_idx' : 'Unnamed: 29'}.items():
        idxes[idx_name] = quantileEncoding(univ,col)


    wb = openpyxl.Workbook()
    ws = wb.active

    for idx in range(len(idxes['living_idx'])): 
        ws.cell(row = idx + 1, column = 1).value =  idxes['gvi_idx'][idx]
        ws.cell(row = idx + 1, column = 2).value =  idxes['sdg_idx'][idx]
        ws.cell(row = idx + 1, column = 3).value =  idxes['living_idx'][idx]
        ws.cell(row = idx + 1, column = 4).value =  idxes['crime_idx'][idx]

        
    wb.save('idx_data.xlsx')
    
    